import asyncio
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout
import pandas as pd
from typing import Dict
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re


class TradingViewScraperOptimized:
    def __init__(self, url: str, headless: bool = False):
        self.url = url
        self.headless = headless
        self.playwright = None
        self.browser = None
        self.page = None
        self.total_rows_loaded = 0

    async def start(self):
        """Initialize browser"""
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(
            headless=self.headless,
            args=['--no-sandbox', '--disable-setuid-sandbox', '--disable-dev-shm-usage']
        )

        context = await self.browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            viewport={'width': 1920, 'height': 1080}
        )

        self.page = await context.new_page()
        self.page.set_default_timeout(60000)

    async def close(self):
        """Properly cleanup all resources"""
        try:
            if self.browser:
                await self.browser.close()
                print("\n🔒 Browser closed")
        except Exception as e:
            print(f"⚠ Error closing browser: {e}")

        try:
            if self.playwright:
                await self.playwright.stop()
                print("🔒 Playwright stopped")
        except Exception as e:
            print(f"⚠ Error stopping playwright: {e}")

    async def load_page(self):
        """Navigate to page"""
        print("Loading page...")
        await self.page.goto(self.url, wait_until='domcontentloaded')
        await asyncio.sleep(3)
        print("✓ Page loaded")

    async def click_tab(self, tab_id: str, tab_name: str) -> bool:
        """Click tab and wait for content"""
        try:
            await self.page.click(f'button[id="{tab_id}"]')
            await asyncio.sleep(1.5)
            print(f"  ✓ Clicked tab: {tab_name}")
            return True
        except Exception as e:
            print(f"  ✗ Error clicking tab {tab_name}: {e}")
            return False

    async def get_current_row_count(self) -> int:
        """Get current number of loaded rows"""
        try:
            count = await self.page.locator('tbody[data-testid="selectable-rows-table-body"] tr.listRow').count()
            return count
        except:
            return 0

    async def load_all_rows(self, is_first_tab: bool = False) -> int:
        """Click Load More until all rows loaded - SMART VERSION"""

        if not is_first_tab:
            current_count = await self.get_current_row_count()
            if current_count >= self.total_rows_loaded * 0.95:
                print(f"  ⚡ Rows already loaded ({current_count} rows) - skipping Load More")
                return 0

        clicks = 0
        max_clicks = 150
        consecutive_failures = 0

        print("  Loading all rows...")

        while clicks < max_clicks:
            try:
                button = self.page.locator('button:has-text("Load More")').first

                if not await button.is_visible(timeout=2000):
                    break

                await button.scroll_into_view_if_needed()
                await button.click()
                clicks += 1
                consecutive_failures = 0

                if clicks % 10 == 0:
                    print(f"    {clicks} clicks...")

                await asyncio.sleep(0.8)

            except PlaywrightTimeout:
                consecutive_failures += 1
                if consecutive_failures >= 2:
                    break
            except Exception:
                break

        final_count = await self.get_current_row_count()
        if final_count > self.total_rows_loaded:
            self.total_rows_loaded = final_count

        print(f"  ✓ Loaded all rows ({clicks} clicks, {final_count} total rows)")
        return clicks

    async def extract_table_data_fast(self, tab_name: str) -> pd.DataFrame:
        """Extract data using JavaScript - WITH TICKER SEPARATION"""
        try:
            start_time = time.time()
            print(f"  Extracting data...")

            # ✅ ENHANCED: Extract ticker symbol and company name separately
            table_data = await self.page.evaluate('''() => {
                const headers = [];
                const rows = [];

                const headerCells = document.querySelectorAll('thead.tableHead-RHkwFEqU th[data-field]');
                headerCells.forEach(cell => {
                    const field = cell.getAttribute('data-field');
                    if (field) headers.push(field);
                });

                const rowElements = document.querySelectorAll('tbody[data-testid="selectable-rows-table-body"] tr.listRow');
                rowElements.forEach(row => {
                    const cells = row.querySelectorAll('td.cell-RLhfr_y4');
                    const rowData = [];

                    cells.forEach((cell, index) => {
                        // ✅ Special handling for ticker column (usually first column)
                        if (index === 0) {
                            // Try to extract ticker symbol and company name separately
                            const symbolElement = cell.querySelector('[class*="tickerNameBox"]');
                            const descElement = cell.querySelector('[class*="tickerDescription"]');

                            if (symbolElement && descElement) {
                                rowData.push(symbolElement.innerText.trim()); // Ticker symbol
                                rowData.push(descElement.innerText.trim());   // Company name
                            } else {
                                // Fallback: try to split by newline or pattern
                                const fullText = cell.innerText.trim();
                                const lines = fullText.split('\\n');
                                if (lines.length >= 2) {
                                    rowData.push(lines[0].trim()); // Symbol
                                    rowData.push(lines.slice(1).join(' ').trim()); // Name
                                } else {
                                    rowData.push(fullText);
                                    rowData.push('');
                                }
                            }
                        } else {
                            rowData.push(cell.innerText.trim());
                        }
                    });

                    if (rowData.length > 0) {
                        rows.push(rowData);
                    }
                });

                return { headers, rows };
            }''')

            elapsed = time.time() - start_time

            headers = table_data['headers']
            rows_data = table_data['rows']

            # ✅ Adjust headers to include "Company Name" after "Ticker"
            if headers and headers[0].lower() in ['ticker', 'symbol']:
                headers_adj = ['Ticker', 'Company Name'] + headers[1:]
            else:
                headers_adj = headers

            print(f"  ✓ Extracted {len(rows_data)} rows × {len(headers_adj)} columns in {elapsed:.1f}s")

            if headers_adj and rows_data:
                max_cols = max(len(row) for row in rows_data) if rows_data else len(headers_adj)

                # Adjust headers to match max columns
                if len(headers_adj) < max_cols:
                    headers_adj = headers_adj + [f'Column_{i}' for i in range(len(headers_adj), max_cols)]
                else:
                    headers_adj = headers_adj[:max_cols]

                # Pad rows
                rows_padded = [row + [''] * (max_cols - len(row)) for row in rows_data]

                df = pd.DataFrame(rows_padded, columns=headers_adj)

                # ✅ Clean up ticker and company name if still combined
                if 'Ticker' in df.columns:
                    df = self.split_ticker_column(df)

                return df

            return pd.DataFrame()

        except Exception as e:
            print(f"  ✗ Error extracting data: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def split_ticker_column(self, df: pd.DataFrame) -> pd.DataFrame:
        """Split ticker column if symbol and name are still combined"""
        if 'Ticker' not in df.columns:
            return df

        # Check if Company Name column already exists and is populated
        if 'Company Name' in df.columns and df['Company Name'].notna().any():
            return df

        # Try to split Ticker column
        def extract_ticker_parts(text):
            if pd.isna(text) or text == '':
                return ('', '')

            # Pattern: "AAPL\nApple Inc." or "AAPL Apple Inc."
            lines = str(text).split('\n')
            if len(lines) >= 2:
                return (lines[0].strip(), ' '.join(lines[1:]).strip())

            # Pattern: "AAPL" (ticker only, usually all caps, short)
            # vs "Apple Inc." (company name, longer, mixed case)
            text = str(text).strip()
            if len(text) <= 5 and text.isupper():
                return (text, '')

            # Try to find pattern: UPPERCASE followed by text
            match = re.match(r'^([A-Z]{1,5})\s+(.+)$', text)
            if match:
                return (match.group(1), match.group(2))

            return (text, '')

        if 'Company Name' not in df.columns:
            df.insert(1, 'Company Name', '')

        ticker_parts = df['Ticker'].apply(extract_ticker_parts)
        df['Ticker'] = ticker_parts.apply(lambda x: x[0])
        df['Company Name'] = ticker_parts.apply(lambda x: x[1])

        return df

    async def scrape_tab(self, tab_id: str, tab_name: str, is_first_tab: bool = False) -> pd.DataFrame:
        """Scrape complete tab"""
        print(f"\n{'=' * 60}")
        print(f"Processing: {tab_name}")
        print(f"{'=' * 60}")

        tab_start = time.time()

        if not await self.click_tab(tab_id, tab_name):
            return pd.DataFrame()

        await self.load_all_rows(is_first_tab=is_first_tab)
        df = await self.extract_table_data_fast(tab_name)

        elapsed = time.time() - tab_start

        if not df.empty:
            print(f"✓ {tab_name}: {len(df)} rows in {elapsed:.1f}s")
        else:
            print(f"✗ {tab_name}: No data")

        return df

    async def scrape_all_tabs(self) -> Dict[str, pd.DataFrame]:
        """Scrape all 9 tabs - OPTIMIZED"""
        tabs = [
            ("overview", "Overview"),
            ("performance", "Performance"),
            ("valuation", "Valuation"),
            ("dividends", "Dividends"),
            ("profitability", "Profitability"),
            ("incomeStatement", "Income Statement"),
            ("balanceSheet", "Balance Sheet"),
            ("cashFlow", "Cash Flow"),
            ("technicals", "Technicals")
        ]

        all_data = {}

        for idx, (tab_id, tab_name) in enumerate(tabs):
            is_first = (idx == 0)
            df = await self.scrape_tab(tab_id, tab_name, is_first_tab=is_first)
            all_data[tab_name] = df
            await asyncio.sleep(0.5)

        return all_data

    def format_excel_professionally(self, filename: str):
        """Apply professional Excel formatting - BEAUTIFUL OUTPUT"""
        print(f"\n{'=' * 60}")
        print("🎨 Applying professional formatting...")
        print(f"{'=' * 60}")

        wb = load_workbook(filename)

        # Define styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark blue
        header_font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')  # White text
        border_style = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  📄 Formatting '{sheet_name}'...")

            # ✅ 1. Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border_style

            # ✅ 2. Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width

            # ✅ 3. Format data cells
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                for col_idx, cell in enumerate(row, start=1):
                    cell.border = border_style

                    if cell.value is not None and cell.value != '':
                        # Detect and format numbers
                        cell_value = str(cell.value).strip()

                        # Skip ticker and company name columns (usually first two)
                        if col_idx <= 2:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                            continue

                        # Format percentages
                        if '%' in cell_value:
                            try:
                                num = float(cell_value.replace('%', '').replace(',', ''))
                                cell.value = num / 100
                                cell.number_format = '0.00%'
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            except:
                                cell.alignment = Alignment(horizontal='right', vertical='center')

                        # Format currency
                        elif '$' in cell_value or cell_value.replace(',', '').replace('.', '').replace('-', '').replace(
                                '+', '').isdigit():
                            try:
                                num = float(cell_value.replace('$', '').replace(',', '').replace('+', ''))
                                cell.value = num

                                # Choose format based on magnitude
                                if abs(num) >= 1_000_000_000:
                                    cell.number_format = '$#,##0.00,,"B"'  # Billions
                                elif abs(num) >= 1_000_000:
                                    cell.number_format = '$#,##0.00,"M"'  # Millions
                                elif abs(num) >= 1000:
                                    cell.number_format = '$#,##0.00,"K"'  # Thousands
                                else:
                                    cell.number_format = '$#,##0.00'

                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            except:
                                # Try as regular number
                                try:
                                    num = float(cell_value.replace(',', '').replace('+', ''))
                                    cell.value = num
                                    cell.number_format = '#,##0.00'
                                    cell.alignment = Alignment(horizontal='right', vertical='center')
                                except:
                                    cell.alignment = Alignment(horizontal='left', vertical='center')

                        # Format regular numbers
                        elif cell_value.replace(',', '').replace('.', '').replace('-', '').replace('+', '').isdigit():
                            try:
                                num = float(cell_value.replace(',', '').replace('+', ''))
                                cell.value = num
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            except:
                                cell.alignment = Alignment(horizontal='left', vertical='center')

                        # Text alignment
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')

            # ✅ 4. Freeze header row
            ws.freeze_panes = 'A2'

            # ✅ 5. Add autofilter
            if ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions

            # ✅ 6. Set row height for header
            ws.row_dimensions[1].height = 25

            print(f"    ✓ {ws.max_row - 1:,} rows formatted")

        wb.save(filename)
        print(f"\n✅ Professional formatting applied!")

    def save_to_excel(self, data: Dict[str, pd.DataFrame], filename: str):
        """Save to Excel with basic structure"""
        print(f"\n{'=' * 60}")
        print("💾 Saving to Excel...")
        print(f"{'=' * 60}")

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            for sheet_name, df in data.items():
                if not df.empty:
                    safe_name = sheet_name[:31]
                    df.to_excel(writer, sheet_name=safe_name, index=False)
                    print(f"  ✓ '{safe_name}': {len(df)} rows")

        # ✅ Apply professional formatting
        self.format_excel_professionally(filename)

    async def run(self, output_file: str = "tradingview_data.xlsx"):
        """Main execution"""
        overall_start = time.time()

        try:
            await self.start()
            await self.load_page()
            data = await self.scrape_all_tabs()
            self.save_to_excel(data, output_file)

            total_time = time.time() - overall_start

            print(f"\n{'=' * 60}")
            print("✅ SCRAPING COMPLETE!")
            print(f"{'=' * 60}")
            total_rows = sum(len(df) for df in data.values())
            print(f"Total rows: {total_rows:,}")
            print(f"Total time: {total_time / 60:.1f} minutes ({total_time:.1f}s)")
            print(f"Average: {total_time / 9:.1f}s per tab")
            print(f"Output: {output_file}")

        except Exception as e:
            print(f"\n✗ Error: {e}")
            import traceback
            traceback.print_exc()
        finally:
            await self.close()


# ===================================================================
# USAGE
# ===================================================================
async def main():
    url = "https://www.tradingview.com/markets/stocks-usa/market-movers-all-stocks/"
    scraper = TradingViewScraperOptimized(url, headless=False)
    await scraper.run("tradingview_all_us_stocks.xlsx")


if __name__ == "__main__":
    asyncio.run(main())